#!/usr/bin/env python3
"""
Scraper unificado: WIPO Madrid Monitor (OMPI) + Marcanet IMPI
Genera un Excel con 3 hojas:
  1. Resultados OMPI
  2. Resultados IMPI (con métricas de similitud)
  3. Análisis de Riesgo y Concesión (criterios combinados)

OMPI usa Playwright (Chromium) para sortear Cloudflare/jqGrid.
"""

import argparse
import json
import os
import re
import sys
import time
import random
import traceback
import warnings
import numpy as np

# ── Playwright browser path (necesario para Render y otros servidores) ────────
# Apunta los browsers al directorio del proyecto en vez de ~/.cache/
# para que persistan entre build y runtime.
_PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))
os.environ.setdefault("PLAYWRIGHT_BROWSERS_PATH", os.path.join(_PROJECT_DIR, ".playwright"))

from datetime import datetime
from pathlib import Path

import pandas as pd
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from bs4 import BeautifulSoup, XMLParsedAsHTMLWarning
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Border, Side, Alignment, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Playwright (OMPI) ────────────────────────────────────────────────────────
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError

import Levenshtein
import jellyfish
from xgboost import XGBClassifier
from sklearn.model_selection import train_test_split
from tqdm import tqdm


warnings.filterwarnings("ignore", category=XMLParsedAsHTMLWarning)

try:
    from rich.console import Console
    from rich.progress import (
        Progress, BarColumn, TextColumn, TimeElapsedColumn,
        TimeRemainingColumn, SpinnerColumn, MofNCompleteColumn,
    )
    from rich.table import Table
    from rich.panel import Panel
    RICH_AVAILABLE = True
except ImportError:
    RICH_AVAILABLE = False

console = Console() if RICH_AVAILABLE else None

# ── Constantes OMPI (Selenium) ────────────────────────────────────────────────
_WIPO_MONITOR_URL    = "https://www3.wipo.int/madrid/monitor/en/"
_WIPO_WAIT_TIMEOUT   = 25   # segundos máximos de espera por elemento
_WIPO_SLEEP_PAGES    = 3    # segundos entre páginas


# ── Helpers comunes ───────────────────────────────────────────────────────────

def default_downloads_path(filename: str) -> str:
    home = Path.home()
    for c in [home / "Downloads", home / "Descargas"]:
        if c.exists():
            return str(c / filename)
    return str(home / filename)


def parse_manual_classes(s: str) -> list:
    s = s.strip()
    if not s:
        raise ValueError("Entrada vacía.")
    tokens = [t.strip() for t in s.split(",") if t.strip()]
    classes = set()
    for t in tokens:
        m = re.fullmatch(r"(\d+)\s*-\s*(\d+)", t)
        if m:
            a, b = int(m.group(1)), int(m.group(2))
            if a > b: a, b = b, a
            for x in range(a, b + 1):
                if 1 <= x <= 45: classes.add(x)
            continue
        if re.fullmatch(r"\d+", t):
            x = int(t)
            if 1 <= x <= 45: classes.add(x)
            continue
        raise ValueError(f"Formato inválido: '{t}'")
    return sorted(classes)


def prompt_menu_and_get_classes() -> list:
    print("\nElige clases a analizar:")
    print("[T]odas  [S]ervicios(35-45)  [P]roductos(1-34)  [M]anual")
    opt = input("Selecciona opción: ").strip().upper()
    while opt not in {"T", "S", "P", "M"}:
        opt = input("Opción inválida. Selecciona T/S/P/M: ").strip().upper()
    if opt == "T": return list(range(1, 46))
    if opt == "S": return list(range(35, 46))
    if opt == "P": return list(range(1, 35))
    while True:
        raw = input("Ingresa clases (ej. 35,41 o 1-5,35-37): ").strip()
        try:
            classes = parse_manual_classes(raw)
            if classes: return classes
            print("No se detectaron clases válidas.")
        except Exception as e:
            print(f"Error: {e}")


def prompt_brand() -> str:
    brand = input("Ingresa la MARCA NUEVA: ").strip()
    while not brand:
        brand = input("La marca no puede ir vacía: ").strip()
    return brand


# ── Modelo ML ─────────────────────────────────────────────────────────────────

def train_model(n_obs=1000):
    base = [w.lower() for w in [
        'Alpha','Beta','Gamma','Delta','Omega','Vertex','Nexus','Pioneer',
        'Quantum','Nova','Apex','Summit','Fusion','Spectrum','Vector',
        'Prime','Edge','Core','Pulse','Echo','Zenith','Horizon','Vista'
    ]]
    suf = [s.lower() for s in [
        'Tech','Tek','Systems','Solutions','Dynamics','Global','Works',
        'Labs','Corp','Soft'
    ]]
    X, y = [], []
    for _ in range(n_obs):
        b1 = random.choice(base) + " " + random.choice(suf)
        b2 = random.choice(base) + " " + random.choice(suf)
        d  = Levenshtein.distance(b1, b2) / max(len(b1), len(b2), 1)
        f2 = jellyfish.jaro_winkler_similarity(b1, b2)
        X.append([d, f2])
        y.append(int(d < 0.4 or f2 > 0.85))
    X_tr, X_te, y_tr, y_te = train_test_split(X, y, test_size=0.2, random_state=42)
    m = XGBClassifier(eval_metric="logloss", verbosity=0)
    m.fit(np.array(X_tr), np.array(y_tr))
    return m


def compute_metrics(marca_nueva: str, marca_existente: str, model) -> dict:
    mn_low = marca_nueva.lower().strip()
    me_low = str(marca_existente).lower().strip()
    d   = Levenshtein.distance(mn_low, me_low) / max(len(mn_low), len(me_low), 1)
    sim = jellyfish.jaro_winkler_similarity(mn_low, me_low)
    prob_conf = model.predict_proba([[d, sim]])[0, 1]
    if mn_low in me_low or mn_low == me_low:
        prob_conf = 1.0
    mn_clean = re.sub(r'[^a-z0-9]', '', mn_low)
    me_clean = re.sub(r'[^a-z0-9]', '', me_low)
    if mn_clean and me_clean and (mn_clean in me_clean or me_clean in mn_clean or mn_clean == me_clean):
        prob_conf = max(prob_conf, 0.9)
    return {
        "Distancia":  round(d, 4),
        "Similitud":  round(sim, 4),
        "Prob_Conf":  round(float(prob_conf), 4),
        "Prob_Conc":  round(float(1 - prob_conf), 4),
    }


# ── OMPI / WIPO Scraper — Playwright ─────────────────────────────────────────

def _selenium_scrape_class(brand: str, niza_class: int, headless: bool = True) -> list:
    """
    Raspa WIPO Madrid Monitor para UNA clase Niza usando Playwright Chromium.
    Devuelve lista de dicts con claves:
        Trademark, Status, Origin, Holder, Req. No, Req. Date, Nice Cl., Vienna Cl., Image URL
    """
    raw_rows = []

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=headless,
            args=[
                "--no-sandbox",
                "--disable-dev-shm-usage",
                "--disable-blink-features=AutomationControlled",
            ],
        )
        context = browser.new_context(
            viewport={"width": 1400, "height": 900},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
        )
        pw_page = context.new_page()
        pw_page.add_init_script(
            "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
        )

        try:
            # ── Paso 1: Abrir página ──────────────────────────────────────
            print(f"   [OMPI] Clase {niza_class}: abriendo Madrid Monitor…")
            pw_page.goto(_WIPO_MONITOR_URL, timeout=_WIPO_WAIT_TIMEOUT * 2 * 1000)
            time.sleep(3)

            # ── Paso 2: Modo avanzado ─────────────────────────────────────
            pw_page.wait_for_selector("#advancedModeLink", timeout=_WIPO_WAIT_TIMEOUT * 1000)
            pw_page.click("#advancedModeLink")
            time.sleep(2)

            # ── Paso 3: Campo Phonetic (marca) ────────────────────────────
            pw_page.wait_for_selector("#BRAND_P_input", timeout=_WIPO_WAIT_TIMEOUT * 1000)
            pw_page.fill("#BRAND_P_input", brand)
            time.sleep(0.5)
            pw_page.click("body")
            time.sleep(0.5)

            # ── Paso 4: Campo Nice (clase Niza) ───────────────────────────
            pw_page.wait_for_selector("#NC_input", timeout=_WIPO_WAIT_TIMEOUT * 1000)
            pw_page.fill("#NC_input", str(niza_class))
            time.sleep(0.5)
            pw_page.click("body")
            time.sleep(0.5)

            # ── Paso 5: Botón Search ──────────────────────────────────────
            pw_page.wait_for_selector("a.searchButton.noPrint", timeout=_WIPO_WAIT_TIMEOUT * 1000)
            pw_page.evaluate("document.querySelector('a.searchButton.noPrint').click()")

            # ── Paso 6: Esperar jqGrid con resultados ─────────────────────
            pw_page.wait_for_selector("div.ui-jqgrid-bdiv", timeout=_WIPO_WAIT_TIMEOUT * 1000)
            pw_page.wait_for_selector("div.ui-jqgrid-bdiv tr.jqgrow", timeout=_WIPO_WAIT_TIMEOUT * 1000)
            time.sleep(2)

            # Total de resultados (informativo)
            try:
                pager_text = pw_page.text_content(".pagerPos")
                total_str  = pager_text.split("/")[-1].strip().replace(",", "")
                total      = int(total_str)
                print(f"   [OMPI] Clase {niza_class}: {total} resultados en servidor")
            except Exception:
                print(f"   [OMPI] Clase {niza_class}: total no disponible")

            # ID del grid (detectado dinámicamente)
            grid_id = pw_page.evaluate("""
                (() => {
                    var grids = document.querySelectorAll('table.ui-jqgrid-btable');
                    return grids.length > 0 ? grids[0].id : 'gridForsearch_pane';
                })()
            """)

            def _clean(val):
                if not val:
                    return ""
                if str(val).strip().startswith("function") or str(val).strip().startswith("<"):
                    return ""
                return re.sub(r'<[^>]+>', '', str(val)).strip()

            def _get_field(row_dict, *keys):
                for k in keys:
                    if k in row_dict:
                        v = _clean(row_dict[k])
                        if v:
                            return v
                return ""

            # ── Paso 7: Iterar páginas ────────────────────────────────────
            page_num = 1
            while True:
                pw_page.wait_for_selector(
                    "div.ui-jqgrid-bdiv tr.jqgrow", timeout=_WIPO_WAIT_TIMEOUT * 1000
                )

                # Datos via API JS de jqGrid (evita formatters HTML)
                rows_data = pw_page.evaluate(
                    f"jQuery('#{grid_id}').jqGrid('getRowData');"
                )

                # URLs de imagen (no están en getRowData)
                img_urls = pw_page.evaluate("""
                    (() => {
                        var result = {};
                        document.querySelectorAll('div.ui-jqgrid-bdiv tr.jqgrow').forEach(function(tr) {
                            var img = tr.querySelector('img');
                            result[tr.id] = img ? img.src : '';
                        });
                        return result;
                    })()
                """)

                # Status desde DOM
                status_map = pw_page.evaluate("""
                    (() => {
                        var result = {};
                        document.querySelectorAll('div.ui-jqgrid-bdiv tr.jqgrow').forEach(function(tr) {
                            var cell = tr.querySelector('td[aria-describedby$="_STATUS"]');
                            var status = '';
                            if (cell) {
                                var img = cell.querySelector('img');
                                if (img) {
                                    status = img.title || img.alt || '';
                                }
                                if (!status) {
                                    status = cell.title || cell.textContent || '';
                                }
                            }
                            result[tr.id] = status.trim();
                        });
                        return result;
                    })()
                """)

                row_ids = pw_page.evaluate("""
                    Array.from(
                        document.querySelectorAll('div.ui-jqgrid-bdiv tr.jqgrow')
                    ).map(function(tr) { return tr.id; })
                """)

                for i, row_dict in enumerate(rows_data):
                    row_id = row_ids[i] if i < len(row_ids) else ""
                    raw_rows.append({
                        "Trademark":  _get_field(row_dict, "BRAND", "BRAND_P", "MARK_P", "trademark", "brandName"),
                        "Image URL":  img_urls.get(row_id, ""),
                        "Status":     status_map.get(row_id, "") or _get_field(row_dict, "STATUS", "status", "ST"),
                        "Origin":     _get_field(row_dict, "ORIGIN", "origin", "OR", "or"),
                        "Holder":     _get_field(row_dict, "HOLDER", "holder", "HOL"),
                        "Req. No":    _get_field(row_dict, "IRN", "REQ_NO", "reqNo", "MADRID_ID", "madrid_id", "id"),
                        "Req. Date":  _get_field(row_dict, "REQ_DATE", "reqDate", "FILING_DATE"),
                        "Nice Cl.":   _get_field(row_dict, "NC", "NICE_CL", "niceCl"),
                        "Vienna Cl.": _get_field(row_dict, "VIENNA_CL", "viennaCl", "VC"),
                    })

                # ── Siguiente página ──────────────────────────────────────
                try:
                    is_last = pw_page.evaluate("""
                        (() => {
                            var btn = document.querySelector('a[aria-label="next page"]');
                            if (!btn) btn = document.querySelector('td[id$="_next"]');
                            if (!btn) return true;
                            return btn.className.indexOf('ui-state-disabled') !== -1
                                || btn.className.indexOf('button-disabled') !== -1;
                        })()
                    """)
                    if is_last:
                        print(f"   [OMPI] Clase {niza_class}: última página ({page_num}).")
                        break

                    pw_page.evaluate("""
                        var btn = document.querySelector('a[aria-label="next page"]');
                        if (!btn) btn = document.querySelector('td[id$="_next"]');
                        if (btn) btn.click();
                    """)
                    time.sleep(_WIPO_SLEEP_PAGES)

                    pw_page.wait_for_selector(
                        "div.ui-jqgrid-bdiv tr.jqgrow", timeout=_WIPO_WAIT_TIMEOUT * 1000
                    )
                    page_num += 1

                except Exception as e:
                    print(f"   [OMPI] Clase {niza_class}: no hay siguiente página ({e}).")
                    break

        except Exception as e:
            tb = traceback.format_exc()
            print(f"   [OMPI] Clase {niza_class}: ERROR — {e}\n{tb}")
            raise  # re-lanzar para que scrape_ompi capture el error real

        finally:
            browser.close()

    return raw_rows


def _map_ompi_row(raw: dict, niza_class: int) -> dict | None:
    """
    Convierte un dict con claves del scraper Selenium al formato interno
    que usan build_excel y los análisis de métricas.
    La columna clave para los análisis es 'Marca' (mapeada desde 'Trademark').
    """
    trademark = raw.get("Trademark", "").strip()
    if not trademark:
        return None  # fila vacía, descartar

    # Siempre usar la clase que se buscó, no las clases del registro OMPI
    nice_cl = str(niza_class)

    return {
        "Clase Niza":  nice_cl,
        "N.°de reg.":  raw.get("Req. No", "").strip(),
        "Marca":       trademark,          # ← columna base de todos los análisis
        "Situación":   raw.get("Status", "").strip(),
        "Titular":     raw.get("Holder", "").strip(),
        "País origen": raw.get("Origin", "").strip(),
        "Origen":      "OMPI",
    }


def _dedupe_ompi_rows(rows: list) -> list:
    seen = set()
    out  = []
    for row in rows:
        key = (
            str(row.get("Clase Niza", "")).strip(),
            str(row.get("N.°de reg.", "")).strip().lower(),
            str(row.get("Marca", "")).strip().lower(),
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(row)
    return out


def scrape_ompi(brand: str, classes: list, headless: bool = True) -> list:
    """
    Punto de entrada OMPI.
    Llama al scraper Playwright por cada clase Niza y devuelve los registros
    en el formato esperado por build_excel (columna 'Marca' ← 'Trademark').
    """
    print("   OMPI: iniciando scraping con Playwright (Chromium)…")
    all_rows = []
    last_error = None

    for cls in tqdm(classes, desc="OMPI clases", unit="clase"):
        try:
            raw_rows = _selenium_scrape_class(brand, cls, headless=headless)
            mapped   = [_map_ompi_row(r, cls) for r in raw_rows]
            mapped   = [r for r in mapped if r is not None]
            mapped   = _dedupe_ompi_rows(mapped)
            print(f"   OMPI clase {cls}: {len(mapped)} registros")
            all_rows.extend(mapped)
        except Exception as e:
            tb = traceback.format_exc()
            last_error = f"OMPI clase {cls}: {e}\n{tb}"
            print(f"   ⚠️  {last_error}")

    all_rows = _dedupe_ompi_rows(all_rows)
    print(f"   OMPI: total {len(all_rows)} registros en {len(classes)} clase(s)")

    # Si no se obtuvo ningún resultado y hubo errores, re-lanzar para que app.py lo muestre
    if not all_rows and last_error:
        raise RuntimeError(last_error)

    return all_rows


# ── IMPI / Marcanet Scraper ───────────────────────────────────────────────────

def crear_session_con_retries(total_retries=3, backoff_factor=1.0) -> requests.Session:
    s = requests.Session()
    retry = Retry(
        total=total_retries,
        backoff_factor=backoff_factor,
        status_forcelist=(502, 503, 504),
        allowed_methods=["HEAD", "GET", "OPTIONS", "POST"]
    )
    adapter = HTTPAdapter(max_retries=retry)
    s.mount("https://", adapter)
    s.mount("http://", adapter)
    return s


def buscar_fonetica_impi(clase: str, marca: str) -> pd.DataFrame:
    session = crear_session_con_retries()
    UI_URL = "https://acervomarcas.impi.gob.mx:8181/marcanet/"
    r1 = session.get(UI_URL, timeout=30)
    r1.raise_for_status()
    m = re.search(r'name="javax\.faces\.ViewState".*?value="([^"]+)"', r1.text)
    if not m:
        raise RuntimeError("No pude extraer javax.faces.ViewState.")
    viewstate = m.group(1)
    payload = {
        "javax.faces.partial.ajax":    "true",
        "javax.faces.source":          "frmBsqFonetica:busquedaId2",
        "javax.faces.partial.execute": "@all",
        "javax.faces.partial.render":  "frmBsqFonetica",
        "frmBsqFonetica":              "frmBsqFonetica",
        "frmBsqFonetica:busquedaId2":  "frmBsqFonetica:busquedaId2",
        "frmBsqFonetica:clases":       clase,
        "frmBsqFonetica:denominacion": marca,
        "javax.faces.ViewState":       viewstate,
    }
    headers = {
        "Content-Type":  "application/x-www-form-urlencoded; charset=UTF-8",
        "Faces-Request": "partial/ajax",
        "User-Agent":    "Mozilla/5.0"
    }
    AJAX_URL = "https://acervomarcas.impi.gob.mx:8181/marcanet/vistas/common/datos/bsqFoneticaCompleta.pgi"
    r2 = session.post(AJAX_URL, data=payload, headers=headers, timeout=30)
    r2.raise_for_status()
    root = BeautifulSoup(r2.text, features="xml")
    upd  = root.find("update", {"id": re.compile(r"^frmBsqFonetica")})
    fragment = upd.text if upd else ""
    soup  = BeautifulSoup(fragment, "html.parser")
    table = soup.find("table")
    if not table:
        return pd.DataFrame()
    rows = table.select("tbody tr")
    datos = [[td.get_text(strip=True) for td in tr.find_all("td")] for tr in rows]
    if not datos:
        return pd.DataFrame()
    max_cols = max(len(r) for r in datos)
    cols = [f"col{i+1}" for i in range(max_cols)]
    norm = [r + [""] * (max_cols - len(r)) for r in datos]
    df   = pd.DataFrame(norm, columns=cols)
    df   = df.loc[:, ["col2", "col4", "col5", "col7"]]
    df.columns = ["TIPO", "Propietario", "No_de_Registro", "Marca_Existente"]
    df   = df[df["No_de_Registro"].str.strip().str.match(r"^\d+$", na=False)]
    df   = df.reset_index(drop=True)
    return df


def scrape_impi(brand: str, classes: list, model) -> list:
    registros = []
    for c in tqdm(classes, desc="IMPI clases", unit="clase"):
        try:
            dfc = buscar_fonetica_impi(str(c), brand)
            if dfc.empty:
                continue
            for _, fila in dfc.iterrows():
                me  = fila['Marca_Existente']
                reg = fila['No_de_Registro']
                metrics = compute_metrics(brand, me, model)
                registros.append({
                    'Clase':           c,
                    'Expediente':      reg,
                    'Marca_Existente': me,
                    **metrics
                })
        except Exception as e:
            print(f"⚠️  Clase {c} IMPI falló: {e}")
    return registros


# ── Criterios de riesgo ───────────────────────────────────────────────────────

def evaluar_criterios(distancia: float, similitud: float, prob_conf: float, prob_conc: float) -> dict:
    """
    Devuelve qué criterios de NO CONCESIÓN aplican y si hay concesión global.
    Criterios de NO CONCESIÓN:
      C1: Distancia  <= 0.30
      C2: Similitud  >= 0.87
      C3: Prob_Conf  >= 0.50
      C4: Prob_Conc  <= 0.89
    Concesión global: si 3 o más criterios NO aplican (≤ 1 criterio activo)
    """
    c1 = distancia  <= 0.30
    c2 = similitud  >= 0.87
    c3 = prob_conf  >= 0.50
    c4 = prob_conc  <= 0.89
    activos = sum([c1, c2, c3, c4])
    hay_concesion = activos <= 1
    criterios_texto = []
    if c1: criterios_texto.append("Distancia≤30%")
    if c2: criterios_texto.append("Similitud≥87%")
    if c3: criterios_texto.append("ProbConf≥50%")
    if c4: criterios_texto.append("ProbConc≤89%")
    return {
        "criterios_activos": activos,
        "criterios_texto":   ", ".join(criterios_texto) if criterios_texto else "Ninguno",
        "hay_concesion":     hay_concesion,
    }


# ── Formateo Excel ────────────────────────────────────────────────────────────

BLUE_DARK   = "000066"
BLUE_MID    = "3366CC"
BLUE_LIGHT  = "99CCFF"
BLACK       = "000000"
WHITE       = "FFFFFF"
HEADER_BG   = "1F3864"
HEADER_BG2  = "2F5496"
HEADER_BG3  = "833C00"
GREEN_BG    = "E2EFDA"
RED_BG      = "FFDDC1"
YELLOW_BG   = "FFFF99"
GRAY_BG     = "F2F2F2"

thin_side   = Side(style="thin",   color="CCCCCC")
thick_side  = Side(style="medium", color="000000")
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)


def header_fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def apply_sheet_header(ws, headers: list, hex_color: str = HEADER_BG):
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill      = header_fill(hex_color)
        cell.font      = Font(bold=True, color=WHITE, name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border


def set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def pct_format(cell):
    cell.number_format = "0.00%"


def apply_row_data(ws, row_idx: int, values: list, pct_cols: list = None, bg: str = None):
    for c_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=c_idx, value=val)
        cell.font      = Font(name="Arial", size=9)
        cell.border    = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        if bg:
            cell.fill = PatternFill(fill_type="solid", fgColor=bg)
        if pct_cols and c_idx in pct_cols:
            pct_format(cell)


def sim_color(val: float) -> str:
    if val >= 0.87: return "FF0000"
    if val >= 0.70: return BLUE_DARK
    if val >= 0.40: return BLUE_MID
    return BLUE_LIGHT


def dist_color(val: float) -> str:
    if val <= 0.30: return "FF0000"
    if val <= 0.50: return "FF9900"
    return "00B050"


# ── Construcción del Excel ────────────────────────────────────────────────────

def build_excel(
    brand: str,
    ompi_rows: list,
    impi_rows: list,
    output_path: str,
    model,
):
    from openpyxl import Workbook

    wb = Workbook()

    # ── Hoja 1: OMPI ──────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "OMPI - Madrid Monitor"
    ws1.freeze_panes = "A2"

    # Enriquecemos OMPI con métricas calculadas desde la columna 'Marca' (← Trademark)
    ompi_enriched = []
    for r in ompi_rows:
        metrics = compute_metrics(brand, r["Marca"], model)
        ompi_enriched.append({**r, **metrics})

    hdrs1 = ["Clase Niza", "N.° Registro", "Marca OMPI", "Situación",
             "Distancia", "Similitud", "Prob_Conf", "Prob_Conc"]
    apply_sheet_header(ws1, hdrs1, HEADER_BG)
    set_col_widths(ws1, {"A":12,"B":18,"C":35,"D":20,"E":12,"F":12,"G":12,"H":12})
    ws1.row_dimensions[1].height = 30

    pct_cols1 = {5, 6, 7, 8}
    for i, r in enumerate(ompi_enriched, start=2):
        bg   = GRAY_BG if i % 2 == 0 else WHITE
        vals = [
            r.get("Clase Niza",""),
            r.get("N.°de reg.",""),
            r.get("Marca",""),         # ← columna Trademark mapeada
            r.get("Situación",""),
            r.get("Distancia", 0),
            r.get("Similitud", 0),
            r.get("Prob_Conf", 0),
            r.get("Prob_Conc", 0),
        ]
        apply_row_data(ws1, i, vals, pct_cols1, bg)
        sim_cell  = ws1.cell(row=i, column=6)
        dist_cell = ws1.cell(row=i, column=5)
        sim_val   = r.get("Similitud", 0)
        dist_val  = r.get("Distancia", 0)
        sim_cell.fill  = PatternFill(fill_type="solid", fgColor=sim_color(sim_val))
        sim_cell.font  = Font(color=WHITE, name="Arial", size=9)
        dist_cell.fill = PatternFill(fill_type="solid", fgColor=dist_color(dist_val))
        dist_cell.font = Font(color=WHITE, name="Arial", size=9)

    # ── Hoja 2: IMPI ──────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("IMPI - Marcanet")
    ws2.freeze_panes = "A2"

    hdrs2 = ["Clase", "Expediente", "Marca Existente", "Distancia",
             "Similitud", "Prob_Conf", "Prob_Conc"]
    apply_sheet_header(ws2, hdrs2, HEADER_BG2)
    set_col_widths(ws2, {"A":10,"B":18,"C":35,"D":12,"E":12,"F":12,"G":12})
    ws2.row_dimensions[1].height = 30

    pct_cols2 = {4, 5, 6, 7}
    for i, r in enumerate(impi_rows, start=2):
        bg   = GRAY_BG if i % 2 == 0 else WHITE
        vals = [
            r.get("Clase",""),
            r.get("Expediente",""),
            r.get("Marca_Existente",""),
            r.get("Distancia", 0),
            r.get("Similitud", 0),
            r.get("Prob_Conf", 0),
            r.get("Prob_Conc", 0),
        ]
        apply_row_data(ws2, i, vals, pct_cols2, bg)
        sim_cell  = ws2.cell(row=i, column=5)
        dist_cell = ws2.cell(row=i, column=4)
        sim_val   = r.get("Similitud", 0)
        dist_val  = r.get("Distancia", 0)
        sim_cell.fill  = PatternFill(fill_type="solid", fgColor=sim_color(sim_val))
        sim_cell.font  = Font(color=WHITE, name="Arial", size=9)
        dist_cell.fill = PatternFill(fill_type="solid", fgColor=dist_color(dist_val))
        dist_cell.font = Font(color=WHITE, name="Arial", size=9)

    # ── Hoja 3: Análisis de Riesgo ────────────────────────────────────────────
    ws3 = wb.create_sheet("Análisis de Riesgo")

    set_col_widths(ws3, {
        "A":10,"B":8,"C":22,"D":35,"E":12,"F":12,"G":12,"H":12,"I":14,"J":35,"K":16
    })

    # Fila 1: Título principal
    ws3.merge_cells("A1:K1")
    title_cell = ws3["A1"]
    title_cell.value     = f"ANÁLISIS DE RIESGO DE NO CONCESIÓN  —  Marca: {brand}"
    title_cell.fill      = header_fill(HEADER_BG3)
    title_cell.font      = Font(bold=True, color=WHITE, name="Arial", size=12)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 28

    # Filas 2-7: Leyenda de criterios
    criterios_info = [
        ("CRITERIOS DE NO CONCESIÓN:", ""),
        ("C1 – Distancia", "≤ 30%"),
        ("C2 – Similitud", "≥ 87%"),
        ("C3 – Prob. Confusión (Prob_Conf)", "≥ 50%"),
        ("C4 – Prob. Concesión (Prob_Conc)", "≤ 89%"),
        ("CONCESIÓN GLOBAL:", "Habrá concesión si 3 de 4 criterios de no concesión NO existen (≤ 1 activo)"),
    ]
    for ri, (label, val) in enumerate(criterios_info, start=2):
        ws3.row_dimensions[ri].height = 16
        c1 = ws3.cell(row=ri, column=1, value=label)
        c1.font      = Font(bold=(ri in {2, 7}), name="Arial", size=9)
        c1.fill      = PatternFill(fill_type="solid", fgColor="FFF2CC")
        c1.alignment = Alignment(vertical="center")
        ws3.merge_cells(start_row=ri, start_column=2, end_row=ri, end_column=11)
        c2 = ws3.cell(row=ri, column=2, value=val)
        c2.font      = Font(name="Arial", size=9, bold=(ri in {2, 7}))
        c2.fill      = PatternFill(fill_type="solid", fgColor="FFF2CC")
        c2.alignment = Alignment(vertical="center")

    ws3.row_dimensions[8].height = 6

    # Fila 9: Headers de la tabla de datos
    data_start = 9
    ws3.freeze_panes = f"A{data_start + 1}"
    hdrs3 = ["Origen", "Clase", "N.° Registro / Expediente", "Marca Conflicto",
             "Distancia", "Similitud", "Prob_Conf", "Prob_Conc",
             "Criterios Activos", "Criterios de Riesgo", "¿Hay Concesión?"]

    for col_idx, hdr in enumerate(hdrs3, start=1):
        cell = ws3.cell(row=data_start, column=col_idx, value=hdr)
        cell.fill      = header_fill(HEADER_BG3)
        cell.font      = Font(bold=True, color=WHITE, name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border
    ws3.row_dimensions[data_start].height = 30

    # Combinar registros OMPI + IMPI para el análisis
    # En ambos casos la métrica se calcula desde la columna 'Marca' (OMPI) / 'Marca_Existente' (IMPI)
    combined = []
    for r in ompi_enriched:
        combined.append({
            "Origen":    "OMPI",
            "Clase":     r.get("Clase Niza",""),
            "Registro":  r.get("N.°de reg.",""),
            "Marca":     r.get("Marca",""),      # ← Trademark mapeado
            "Distancia": r.get("Distancia", 0),
            "Similitud": r.get("Similitud", 0),
            "Prob_Conf": r.get("Prob_Conf", 0),
            "Prob_Conc": r.get("Prob_Conc", 0),
        })
    for r in impi_rows:
        combined.append({
            "Origen":    "IMPI",
            "Clase":     r.get("Clase",""),
            "Registro":  r.get("Expediente",""),
            "Marca":     r.get("Marca_Existente",""),
            "Distancia": r.get("Distancia", 0),
            "Similitud": r.get("Similitud", 0),
            "Prob_Conf": r.get("Prob_Conf", 0),
            "Prob_Conc": r.get("Prob_Conc", 0),
        })

    combined.sort(key=lambda x: x["Similitud"], reverse=True)

    marcas_conflicto = []
    pct_cols3 = {5, 6, 7, 8}
    for i, r in enumerate(combined, start=data_start + 1):
        ev       = evaluar_criterios(r["Distancia"], r["Similitud"], r["Prob_Conf"], r["Prob_Conc"])
        hay_conc = ev["hay_concesion"]
        bg       = GREEN_BG if hay_conc else RED_BG

        vals = [
            r["Origen"],
            r["Clase"],
            r["Registro"],
            r["Marca"],
            r["Distancia"],
            r["Similitud"],
            r["Prob_Conf"],
            r["Prob_Conc"],
            ev["criterios_activos"],
            ev["criterios_texto"],
            "✔ SÍ" if hay_conc else "✘ NO",
        ]
        apply_row_data(ws3, i, vals, pct_cols3, bg)

        sim_cell  = ws3.cell(row=i, column=6)
        dist_cell = ws3.cell(row=i, column=5)
        sim_cell.fill  = PatternFill(fill_type="solid", fgColor=sim_color(r["Similitud"]))
        sim_cell.font  = Font(color=WHITE, name="Arial", size=9)
        dist_cell.fill = PatternFill(fill_type="solid", fgColor=dist_color(r["Distancia"]))
        dist_cell.font = Font(color=WHITE, name="Arial", size=9)

        conc_cell = ws3.cell(row=i, column=11)
        if hay_conc:
            conc_cell.font = Font(bold=True, color="006100", name="Arial", size=9)
            conc_cell.fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
        else:
            conc_cell.font = Font(bold=True, color="9C0006", name="Arial", size=9)
            conc_cell.fill = PatternFill(fill_type="solid", fgColor="FFC7CE")

        if not hay_conc:
            marcas_conflicto.append({
                "Marca":    r["Marca"],
                "Registro": r["Registro"],
                "Origen":   r["Origen"],
                "Clase":    r["Clase"],
            })

    # Bloque resumen global
    final_row = data_start + len(combined) + 3

    if not combined:
        resultado_global      = "SIN DATOS"
        hay_concesion_global  = True
    else:
        hay_concesion_global = all(
            evaluar_criterios(r["Distancia"], r["Similitud"], r["Prob_Conf"], r["Prob_Conc"])["hay_concesion"]
            for r in combined
        )
        resultado_global = "SÍ" if hay_concesion_global else "NO"

    ws3.merge_cells(f"A{final_row}:K{final_row}")
    res_cell = ws3[f"A{final_row}"]
    if hay_concesion_global:
        res_cell.value = f"RESULTADO GLOBAL: ✔ SÍ HABRÁ CONCESIÓN DE LA MARCA «{brand}»"
        res_cell.fill  = header_fill("C6EFCE")
        res_cell.font  = Font(bold=True, color="006100", name="Arial", size=13)
    else:
        res_cell.value = f"RESULTADO GLOBAL: ✘ NO HABRÁ CONCESIÓN DE LA MARCA «{brand}»  —  EXISTEN MARCAS EN CONFLICTO"
        res_cell.fill  = header_fill("FFC7CE")
        res_cell.font  = Font(bold=True, color="9C0006", name="Arial", size=13)
    res_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[final_row].height = 30

    # Tabla de marcas en conflicto
    if marcas_conflicto:
        from collections import defaultdict
        clases_por_marca = defaultdict(set)
        for mc in marcas_conflicto:
            key = (mc["Marca"], mc["Registro"], mc["Origen"])
            clases_por_marca[key].add(mc["Clase"])

        seen_keys       = set()
        unique_conflicto = []
        for mc in marcas_conflicto:
            key = (mc["Marca"], mc["Registro"], mc["Origen"])
            if key not in seen_keys:
                seen_keys.add(key)
                unique_conflicto.append(mc)

        conflict_start = final_row + 2
        ws3.merge_cells(f"A{conflict_start}:D{conflict_start}")
        lbl = ws3.cell(row=conflict_start, column=1, value="MARCAS EN CONFLICTO:")
        lbl.font      = Font(bold=True, color=WHITE, name="Arial", size=11)
        lbl.fill      = header_fill("9C0006")
        lbl.alignment = Alignment(horizontal="center", vertical="center")
        ws3.row_dimensions[conflict_start].height = 22

        hdr_row = conflict_start + 1
        for ci, h in enumerate(["Marca", "No. Registro", "Origen (OMPI/IMPI)", "Clases Registradas"], start=1):
            cell = ws3.cell(row=hdr_row, column=ci, value=h)
            cell.fill      = header_fill("FF0000")
            cell.font      = Font(bold=True, color=WHITE, name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = thin_border
        ws3.row_dimensions[hdr_row].height = 22

        for di, mc in enumerate(unique_conflicto, start=hdr_row + 1):
            bg        = RED_BG if di % 2 == 0 else "FFE8E8"
            key       = (mc["Marca"], mc["Registro"], mc["Origen"])
            clases_str = ", ".join(str(c) for c in sorted(clases_por_marca[key]))
            for ci, val in enumerate([mc["Marca"], mc["Registro"], mc["Origen"], clases_str], start=1):
                cell = ws3.cell(row=di, column=ci, value=val)
                cell.font      = Font(name="Arial", size=9)
                cell.fill      = PatternFill(fill_type="solid", fgColor=bg)
                cell.border    = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=(ci == 4))
            ws3.row_dimensions[di].height = 16

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Scraper unificado OMPI + IMPI")
    ts          = datetime.now().strftime("%Y%m%d_%H%M%S")
    default_out = default_downloads_path(f"analisis_marcas_{ts}.xlsx")
    parser.add_argument("--out",        default=default_out, help="Archivo de salida .xlsx")
    parser.add_argument("--headed",     action="store_true", help="Abrir navegador visible (no headless)")
    parser.add_argument("--skip-ompi",  action="store_true", help="Omitir scraping OMPI")
    parser.add_argument("--skip-impi",  action="store_true", help="Omitir scraping IMPI")
    args = parser.parse_args()

    print("=" * 60)
    print("  ANALIZADOR DE MARCAS — OMPI (Madrid) + IMPI (Marcanet)")
    print("=" * 60)

    classes = prompt_menu_and_get_classes()
    brand   = prompt_brand()

    print("\n🔧 Entrenando modelo de similitud…")
    model = train_model(1000)
    print("✅ Modelo listo.\n")

    ompi_rows = []
    impi_rows = []

    if not args.skip_ompi:
        print("🌐 Iniciando scraping OMPI / Madrid Monitor (Selenium)…")
        try:
            ompi_rows = scrape_ompi(brand, classes, headless=not args.headed)
            print(f"   → {len(ompi_rows)} registros encontrados en OMPI.")
        except Exception as e:
            print(f"⚠️  Error en OMPI: {e}")

    if not args.skip_impi:
        print("\n🌐 Iniciando scraping IMPI / Marcanet…")
        impi_rows = scrape_impi(brand, classes, model)
        print(f"   → {len(impi_rows)} registros encontrados en IMPI.")

    print(f"\n📊 Generando Excel: {args.out}")
    build_excel(brand, ompi_rows, impi_rows, args.out, model)
    print(f"✅ Archivo guardado: {args.out}\n")


if __name__ == "__main__":
    main()